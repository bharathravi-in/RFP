"""Compliance routes for managing RFP requirement compliance items."""
from flask import Blueprint, request, jsonify, Response
from flask_jwt_extended import jwt_required, get_jwt_identity
from ..extensions import db
from ..models import ComplianceItem, Project, User, RFPSection

bp = Blueprint('compliance', __name__)


@bp.route('/projects/<int:project_id>/compliance', methods=['GET'])
@jwt_required()
def list_compliance_items(project_id):
    """List all compliance items for a project."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404
    
    if project.organization_id != user.organization_id:
        return jsonify({'error': 'Access denied'}), 403
    
    # Get query parameters for filtering
    category = request.args.get('category')
    status = request.args.get('status')
    
    query = ComplianceItem.query.filter_by(project_id=project_id)
    
    if category:
        query = query.filter(ComplianceItem.category == category)
    if status:
        query = query.filter(ComplianceItem.compliance_status == status)
    
    items = query.order_by(ComplianceItem.order, ComplianceItem.created_at).all()
    
    # Get compliance stats
    total = len(items)
    stats = {
        'total': total,
        'compliant': sum(1 for i in items if i.compliance_status == 'compliant'),
        'partial': sum(1 for i in items if i.compliance_status == 'partial'),
        'non_compliant': sum(1 for i in items if i.compliance_status == 'non_compliant'),
        'not_applicable': sum(1 for i in items if i.compliance_status == 'not_applicable'),
        'pending': sum(1 for i in items if i.compliance_status == 'pending'),
    }
    
    return jsonify({
        'items': [item.to_dict(include_section=True) for item in items],
        'stats': stats,
    }), 200


@bp.route('/projects/<int:project_id>/compliance', methods=['POST'])
@jwt_required()
def create_compliance_item(project_id):
    """Create a new compliance item."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404
    
    if project.organization_id != user.organization_id:
        return jsonify({'error': 'Access denied'}), 403
    
    data = request.get_json()
    if not data or not data.get('requirement_text'):
        return jsonify({'error': 'Requirement text is required'}), 400
    
    # Get max order for new item
    max_order = db.session.query(db.func.max(ComplianceItem.order))\
        .filter_by(project_id=project_id).scalar() or 0
    
    item = ComplianceItem(
        project_id=project_id,
        requirement_id=data.get('requirement_id'),
        requirement_text=data['requirement_text'],
        source=data.get('source'),
        category=data.get('category'),
        compliance_status=data.get('compliance_status', 'pending'),
        section_id=data.get('section_id'),
        response_summary=data.get('response_summary'),
        notes=data.get('notes'),
        priority=data.get('priority', 'normal'),
        order=max_order + 1,
    )
    
    db.session.add(item)
    db.session.commit()
    
    return jsonify({
        'message': 'Compliance item created',
        'item': item.to_dict(include_section=True),
    }), 201


@bp.route('/compliance/<int:item_id>', methods=['GET'])
@jwt_required()
def get_compliance_item(item_id):
    """Get a single compliance item."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    item = ComplianceItem.query.get(item_id)
    if not item:
        return jsonify({'error': 'Compliance item not found'}), 404
    
    if item.project.organization_id != user.organization_id:
        return jsonify({'error': 'Access denied'}), 403
    
    return jsonify({'item': item.to_dict(include_section=True)}), 200


@bp.route('/compliance/<int:item_id>', methods=['PUT'])
@jwt_required()
def update_compliance_item(item_id):
    """Update a compliance item."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    item = ComplianceItem.query.get(item_id)
    if not item:
        return jsonify({'error': 'Compliance item not found'}), 404
    
    if item.project.organization_id != user.organization_id:
        return jsonify({'error': 'Access denied'}), 403
    
    data = request.get_json()
    
    # Update allowed fields
    if 'requirement_id' in data:
        item.requirement_id = data['requirement_id']
    if 'requirement_text' in data:
        item.requirement_text = data['requirement_text']
    if 'source' in data:
        item.source = data['source']
    if 'category' in data:
        item.category = data['category']
    if 'compliance_status' in data:
        item.compliance_status = data['compliance_status']
    if 'section_id' in data:
        item.section_id = data['section_id']
    if 'response_summary' in data:
        item.response_summary = data['response_summary']
    if 'notes' in data:
        item.notes = data['notes']
    if 'priority' in data:
        item.priority = data['priority']
    if 'order' in data:
        item.order = data['order']
    
    db.session.commit()
    
    return jsonify({
        'message': 'Compliance item updated',
        'item': item.to_dict(include_section=True),
    }), 200


@bp.route('/compliance/<int:item_id>', methods=['DELETE'])
@jwt_required()
def delete_compliance_item(item_id):
    """Delete a compliance item."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    item = ComplianceItem.query.get(item_id)
    if not item:
        return jsonify({'error': 'Compliance item not found'}), 404
    
    if item.project.organization_id != user.organization_id:
        return jsonify({'error': 'Access denied'}), 403
    
    db.session.delete(item)
    db.session.commit()
    
    return jsonify({'message': 'Compliance item deleted'}), 200


@bp.route('/projects/<int:project_id>/compliance/bulk', methods=['POST'])
@jwt_required()
def bulk_create_compliance_items(project_id):
    """Bulk create compliance items (for importing from RFP analysis)."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404
    
    if project.organization_id != user.organization_id:
        return jsonify({'error': 'Access denied'}), 403
    
    data = request.get_json()
    items_data = data.get('items', [])
    
    if not items_data:
        return jsonify({'error': 'No items provided'}), 400
    
    # Get max order
    max_order = db.session.query(db.func.max(ComplianceItem.order))\
        .filter_by(project_id=project_id).scalar() or 0
    
    created_items = []
    for i, item_data in enumerate(items_data):
        if not item_data.get('requirement_text'):
            continue
            
        item = ComplianceItem(
            project_id=project_id,
            requirement_id=item_data.get('requirement_id'),
            requirement_text=item_data['requirement_text'],
            source=item_data.get('source'),
            category=item_data.get('category'),
            compliance_status=item_data.get('compliance_status', 'pending'),
            priority=item_data.get('priority', 'normal'),
            order=max_order + i + 1,
        )
        db.session.add(item)
        created_items.append(item)
    
    db.session.commit()
    
    return jsonify({
        'message': f'Created {len(created_items)} compliance items',
        'count': len(created_items),
    }), 201


@bp.route('/projects/<int:project_id>/compliance/export', methods=['GET'])
@jwt_required()
def export_compliance_matrix(project_id):
    """Export compliance matrix to Excel."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404
    
    if project.organization_id != user.organization_id:
        return jsonify({'error': 'Access denied'}), 403
    
    items = ComplianceItem.query.filter_by(project_id=project_id)\
        .order_by(ComplianceItem.order, ComplianceItem.created_at).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"
    
    # Styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Status colors
    status_colors = {
        'compliant': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        'partial': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        'non_compliant': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        'not_applicable': PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
        'pending': PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid"),
    }
    
    # Headers
    headers = ['Req ID', 'Requirement', 'Category', 'Status', 'Section', 'Response Summary', 'Notes', 'Priority']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 10
    
    # Data rows
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.requirement_id or '').border = thin_border
        ws.cell(row=row, column=2, value=item.requirement_text).border = thin_border
        ws.cell(row=row, column=3, value=item.category or '').border = thin_border
        
        status_cell = ws.cell(row=row, column=4, value=item.compliance_status.replace('_', ' ').title())
        status_cell.border = thin_border
        status_cell.fill = status_colors.get(item.compliance_status, PatternFill())
        
        ws.cell(row=row, column=5, value=item.section.title if item.section else '').border = thin_border
        ws.cell(row=row, column=6, value=item.response_summary or '').border = thin_border
        ws.cell(row=row, column=7, value=item.notes or '').border = thin_border
        ws.cell(row=row, column=8, value=item.priority or 'normal').border = thin_border
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"{project.name.replace(' ', '_')}_Compliance_Matrix.xlsx"
    
    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@bp.route('/projects/<int:project_id>/compliance/extract', methods=['POST'])
@jwt_required()
def extract_requirements_from_documents(project_id):
    """
    Extract requirements from project RFP documents using AI.
    Auto-populates the compliance matrix with identified requirements.
    """
    import json
    import re
    import tempfile
    import os
    import logging
    from ..models import Document
    from ..services.document_service import DocumentService
    
    logger = logging.getLogger(__name__)
    
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404
    
    if project.organization_id != user.organization_id:
        return jsonify({'error': 'Access denied'}), 403
    
    # Get all documents for this project
    documents = Document.query.filter_by(project_id=project_id).all()
    if not documents:
        return jsonify({'error': 'No documents found for this project'}), 400
    
    # Extract text from all documents
    all_text = []
    doc_sources = []
    
    for doc in documents:
        try:
            # First try to use already extracted text
            if doc.extracted_text:
                all_text.append(f"=== From: {doc.filename} ===\n{doc.extracted_text}")
                doc_sources.append(doc.filename)
                continue
            
            # Fall back to extracting from binary file_data
            if not doc.file_data:
                continue
            
            # Create temp file for extraction
            temp_dir = tempfile.mkdtemp()
            ext = doc.file_type or 'pdf'
            temp_path = os.path.join(temp_dir, f'doc.{ext}')
            
            with open(temp_path, 'wb') as f:
                f.write(doc.file_data)
            
            text = DocumentService.extract_text(temp_path, ext)
            if text:
                all_text.append(f"=== From: {doc.filename} ===\n{text}")
                doc_sources.append(doc.filename)
            
            # Cleanup
            os.remove(temp_path)
            os.rmdir(temp_dir)
            
        except Exception as e:
            logger.error(f"Failed to extract text from {doc.filename}: {e}")
            continue
    
    if not all_text:
        return jsonify({'error': 'Could not extract text from any documents'}), 400
    
    combined_text = "\n\n".join(all_text)
    
    # Limit text length for API
    max_chars = 50000
    if len(combined_text) > max_chars:
        combined_text = combined_text[:max_chars] + "\n\n[Content truncated for analysis...]"
    
    # Use AI to extract requirements
    try:
        from flask import current_app
        
        # Try dynamic LLM provider first
        org_id = user.organization_id
        model = None
        
        try:
            from app.services.llm_service_helper import get_llm_provider
            model = get_llm_provider(org_id, 'compliance_extractor')
            if model:
                logger.info(f"Compliance extractor using dynamic provider: {model.provider_name}")
        except Exception as e:
            logger.warning(f"Could not load dynamic LLM: {e}")
        
        # Fallback to legacy Google
        if not model:
            import google.generativeai as genai
            api_key = (
                current_app.config.get('GOOGLE_API_KEY') or 
                current_app.config.get('GEMINI_API_KEY') or 
                os.environ.get('GOOGLE_API_KEY') or 
                os.environ.get('GEMINI_API_KEY')
            )
            if not api_key:
                return jsonify({'error': 'AI service not configured'}), 500
            
            genai.configure(api_key=api_key)
            model_name = os.environ.get('GOOGLE_MODEL', 'gemini-2.0-flash')
            model = genai.GenerativeModel(model_name)
        
        prompt = f"""Analyze this RFP (Request for Proposal) document and extract ALL specific requirements that a vendor must comply with.

For each requirement, provide:
1. requirement_id: A unique identifier (e.g., "REQ-001", "REQ-002")
2. requirement_text: The exact requirement text or a clear summary
3. category: One of: technical, security, compliance, legal, pricing, operational, support
4. priority: high, normal, or low based on language used (must, shall = high; should = normal; may = low)
5. source: The section or page where this requirement appears

Focus on:
- Mandatory requirements (must, shall, required)
- Compliance criteria
- Technical specifications
- Security requirements
- Pricing/commercial terms
- Deliverable requirements
- Timeline requirements

Return ONLY a valid JSON array with objects having these fields. No markdown, no explanations.
Example format:
[
  {{"requirement_id": "REQ-001", "requirement_text": "Vendor must provide 24/7 support", "category": "support", "priority": "high", "source": "Section 5.2"}}
]

RFP Document Content:
{combined_text}"""

        response = model.generate_content(prompt)
        response_text = response.text.strip()
        
        # Clean markdown code blocks if present
        if response_text.startswith('```'):
            response_text = re.sub(r'^```(?:json)?\n?', '', response_text)
            response_text = re.sub(r'\n?```$', '', response_text)
        
        requirements = json.loads(response_text)
        
        if not isinstance(requirements, list):
            requirements = [requirements]
        
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse AI response: {e}")
        return jsonify({'error': 'Failed to parse AI analysis response'}), 500
    except Exception as e:
        logger.error(f"AI requirement extraction failed: {e}")
        return jsonify({'error': f'AI analysis failed: {str(e)}'}), 500
    
    # Check for existing requirements to avoid duplicates
    existing_items = ComplianceItem.query.filter_by(project_id=project_id).all()
    existing_texts = {item.requirement_text.lower().strip() for item in existing_items}
    
    # Get max order
    max_order = db.session.query(db.func.max(ComplianceItem.order))\
        .filter_by(project_id=project_id).scalar() or 0
    
    # Create compliance items
    created_items = []
    skipped = 0
    
    for i, req in enumerate(requirements):
        req_text = req.get('requirement_text', '').strip()
        if not req_text:
            continue
        
        # Check for duplicates
        if req_text.lower().strip() in existing_texts:
            skipped += 1
            continue
        
        item = ComplianceItem(
            project_id=project_id,
            requirement_id=req.get('requirement_id'),
            requirement_text=req_text,
            source=req.get('source', ', '.join(doc_sources[:3])),
            category=req.get('category', 'general'),
            compliance_status='pending',
            priority=req.get('priority', 'normal'),
            order=max_order + i + 1,
        )
        db.session.add(item)
        created_items.append(item)
        existing_texts.add(req_text.lower().strip())
    
    db.session.commit()
    
    return jsonify({
        'message': f'Extracted {len(created_items)} requirements from {len(documents)} documents',
        'extracted_count': len(created_items),
        'skipped_duplicates': skipped,
        'documents_analyzed': len(doc_sources),
        'items': [item.to_dict() for item in created_items],
    }), 201
