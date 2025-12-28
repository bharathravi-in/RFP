"""
Batch Export Service

Provides export functionality for answers, knowledge base, and proposals.
"""
import io
import csv
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)


def export_answers_to_csv(answers: List[Dict]) -> bytes:
    """
    Export answers to CSV format.
    
    Args:
        answers: List of answer dictionaries
        
    Returns:
        CSV file as bytes
    """
    output = io.StringIO()
    
    fieldnames = [
        'question_id', 'question_text', 'answer_text', 
        'confidence', 'status', 'category', 'created_at'
    ]
    
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    for answer in answers:
        writer.writerow({
            'question_id': answer.get('question_id', ''),
            'question_text': answer.get('question_text', ''),
            'answer_text': answer.get('answer_text', ''),
            'confidence': answer.get('confidence', ''),
            'status': answer.get('status', ''),
            'category': answer.get('category', ''),
            'created_at': answer.get('created_at', '')
        })
    
    return output.getvalue().encode('utf-8')


def export_answers_to_excel(answers: List[Dict]) -> bytes:
    """
    Export answers to Excel format.
    
    Args:
        answers: List of answer dictionaries
        
    Returns:
        Excel file as bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Answers"
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Headers
    headers = ['Question ID', 'Question', 'Answer', 'Confidence', 'Status', 'Category', 'Created At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Data
    for row, answer in enumerate(answers, 2):
        ws.cell(row=row, column=1, value=answer.get('question_id', ''))
        ws.cell(row=row, column=2, value=answer.get('question_text', ''))
        ws.cell(row=row, column=3, value=answer.get('answer_text', ''))
        ws.cell(row=row, column=4, value=answer.get('confidence', ''))
        ws.cell(row=row, column=5, value=answer.get('status', ''))
        ws.cell(row=row, column=6, value=answer.get('category', ''))
        ws.cell(row=row, column=7, value=answer.get('created_at', ''))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def export_knowledge_base_to_json(items: List[Dict]) -> bytes:
    """
    Export knowledge base items to JSON.
    
    Args:
        items: List of knowledge base items
        
    Returns:
        JSON file as bytes
    """
    export_data = {
        'export_date': datetime.utcnow().isoformat(),
        'item_count': len(items),
        'items': items
    }
    
    return json.dumps(export_data, indent=2, default=str).encode('utf-8')


def import_answers_from_csv(file_content: bytes) -> List[Dict]:
    """
    Import answers from CSV file.
    
    Args:
        file_content: CSV file content as bytes
        
    Returns:
        List of parsed answer dictionaries
    """
    content = file_content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    
    answers = []
    for row in reader:
        answers.append({
            'question_text': row.get('question_text', row.get('question', '')),
            'answer_text': row.get('answer_text', row.get('answer', '')),
            'category': row.get('category', 'general'),
            'confidence': float(row.get('confidence', 0.7)) if row.get('confidence') else 0.7
        })
    
    return answers


def import_answers_from_excel(file_content: bytes) -> List[Dict]:
    """
    Import answers from Excel file.
    
    Args:
        file_content: Excel file content as bytes
        
    Returns:
        List of parsed answer dictionaries
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")
    
    wb = openpyxl.load_workbook(io.BytesIO(file_content))
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value.lower().replace(' ', '_') if cell.value else f'col_{i}' 
               for i, cell in enumerate(ws[1])]
    
    answers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        
        answers.append({
            'question_text': row_dict.get('question_text') or row_dict.get('question', ''),
            'answer_text': row_dict.get('answer_text') or row_dict.get('answer', ''),
            'category': row_dict.get('category', 'general'),
            'confidence': float(row_dict.get('confidence', 0.7)) if row_dict.get('confidence') else 0.7
        })
    
    return answers


class BatchExportService:
    """Service for batch export operations."""
    
    @staticmethod
    def export_project_answers(project_id: int, format: str = 'csv') -> tuple:
        """
        Export all answers for a project.
        
        Args:
            project_id: Project ID
            format: Export format ('csv', 'excel', 'json')
            
        Returns:
            Tuple of (file_bytes, filename, content_type)
        """
        from app.models import Answer, Question
        
        # Get answers with questions
        answers = Answer.query.join(Question).filter(
            Question.project_id == project_id
        ).all()
        
        answer_dicts = []
        for answer in answers:
            answer_dicts.append({
                'question_id': answer.question_id,
                'question_text': answer.question.question_text if answer.question else '',
                'answer_text': answer.answer_text,
                'confidence': answer.confidence,
                'status': answer.status,
                'category': answer.question.category if answer.question else '',
                'created_at': answer.created_at.isoformat() if answer.created_at else ''
            })
        
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        
        if format == 'excel':
            return (
                export_answers_to_excel(answer_dicts),
                f'answers_project_{project_id}_{timestamp}.xlsx',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        elif format == 'json':
            return (
                json.dumps(answer_dicts, indent=2, default=str).encode('utf-8'),
                f'answers_project_{project_id}_{timestamp}.json',
                'application/json'
            )
        else:  # csv
            return (
                export_answers_to_csv(answer_dicts),
                f'answers_project_{project_id}_{timestamp}.csv',
                'text/csv'
            )
    
    @staticmethod
    def export_answer_library(org_id: int, format: str = 'csv') -> tuple:
        """
        Export answer library for an organization.
        
        Args:
            org_id: Organization ID
            format: Export format
            
        Returns:
            Tuple of (file_bytes, filename, content_type)
        """
        from app.models import AnswerLibraryItem
        
        items = AnswerLibraryItem.query.filter_by(org_id=org_id).all()
        
        item_dicts = [{
            'question_text': item.question_text,
            'answer_text': item.answer_text,
            'category': item.category,
            'tags': ','.join(item.tags) if item.tags else '',
            'usage_count': item.usage_count,
            'created_at': item.created_at.isoformat() if item.created_at else ''
        } for item in items]
        
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        
        if format == 'excel':
            return (
                export_answers_to_excel(item_dicts),
                f'answer_library_{timestamp}.xlsx',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return (
                export_answers_to_csv(item_dicts),
                f'answer_library_{timestamp}.csv',
                'text/csv'
            )
