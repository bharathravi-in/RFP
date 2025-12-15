"""Document processing service for parsing uploaded files."""
import os
from typing import Optional


class DocumentService:
    """Handle document parsing and text extraction."""
    
    @staticmethod
    def extract_text_from_pdf(file_path: str) -> str:
        """Extract text content from PDF file."""
        try:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
            return '\n\n'.join(text_parts)
        except Exception as e:
            raise ValueError(f"Failed to extract text from PDF: {e}")
    
    @staticmethod
    def extract_text_from_docx(file_path: str) -> str:
        """Extract text content from DOCX file."""
        try:
            from docx import Document
            doc = Document(file_path)
            return '\n\n'.join([para.text for para in doc.paragraphs if para.text])
        except Exception as e:
            raise ValueError(f"Failed to extract text from DOCX: {e}")
    
    @staticmethod
    def extract_text_from_xlsx(file_path: str) -> str:
        """Extract text content from XLSX file."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            text_parts = []
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' | '.join(str(cell) for cell in row if cell)
                    if row_text:
                        text_parts.append(row_text)
            return '\n'.join(text_parts)
        except Exception as e:
            raise ValueError(f"Failed to extract text from XLSX: {e}")
    
    @classmethod
    def extract_text(cls, file_path: str, file_type: str) -> str:
        """Extract text from document based on file type."""
        extractors = {
            'pdf': cls.extract_text_from_pdf,
            'docx': cls.extract_text_from_docx,
            'xlsx': cls.extract_text_from_xlsx,
            'xls': cls.extract_text_from_xlsx,
            'doc': cls.extract_text_from_docx,
        }
        
        extractor = extractors.get(file_type.lower())
        if not extractor:
            raise ValueError(f"Unsupported file type: {file_type}")
        
        return extractor(file_path)
    
    @staticmethod
    def get_file_metadata(file_path: str) -> dict:
        """Get metadata about a file."""
        stat = os.stat(file_path)
        return {
            'size': stat.st_size,
            'modified': stat.st_mtime,
        }
